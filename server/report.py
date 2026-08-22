"""月度收支 Excel 报表生成（省账通能力）
基于 openpyxl 输出：收支汇总 + 分类明细 + 交易流水 三个工作表。
"""
import os
from datetime import date
from pathlib import Path

import db
from config import DATA_DIR

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:  # pragma: no cover
    OPENPYXL_OK = False


def get_monthly_report(year: int | None = None, month: int | None = None,
                       out_dir: str | None = None) -> dict:
    """生成 {year}年{month}月 收支报表 xlsx，返回文件路径"""
    if not OPENPYXL_OK:
        return {"error": "缺少依赖 openpyxl，请执行 pip install openpyxl"}

    today = date.today()
    year = year or today.year
    month = month or today.month

    summary = db.monthly_summary(year, month)
    txns = db.list_transactions(year, month, limit=1000)

    wb = openpyxl.Workbook()

    title_font = Font(size=16, bold=True)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="B4532A")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------- Sheet1 收支汇总 ----------
    ws = wb.active
    ws.title = "收支汇总"
    ws.append([f"巷子里的AI掌柜 - {year}年{month}月收支报表"])
    ws.merge_cells("A1:C1")
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["项目", "金额（元）", "笔数"])
    for cell in ws[3]:
        cell.font = head_font
        cell.fill = head_fill
        cell.border = border
    ws.append(["收入合计", summary["income"], summary["income_cnt"]])
    ws.append(["支出合计", summary["expense"], summary["expense_cnt"]])
    ws.append(["结余", summary["balance"], ""])
    for row in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
    ws.append([])
    ws.append(["注：收入/支出/结余单位均为元。结余 = 收入 - 支出。"])
    ws["A8"].font = Font(size=9, color="888888")

    # ---------- Sheet2 分类明细 ----------
    ws2 = wb.create_sheet("分类明细")
    ws2.append([f"{year}年{month}月 分类收支明细"])
    ws2.merge_cells("A1:E1")
    ws2["A1"].font = title_font
    ws2.append([])
    ws2.append(["分类", "大白话", "类型", "金额（元）", "笔数"])
    for cell in ws2[3]:
        cell.font = head_font
        cell.fill = head_fill
    for cat in summary["categories"]:
        label = "收入" if cat["trans_type"] == "income" else "支出"
        ws2.append([cat["category"], cat["friendly"], label, cat["total"], cat["cnt"]])
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    # ---------- Sheet3 交易流水 ----------
    ws3 = wb.create_sheet("交易流水")
    ws3.append([f"{year}年{month}月交易流水（共 {len(txns)} 笔）"])
    ws3.merge_cells("A1:H1")
    ws3["A1"].font = title_font
    ws3.append([])
    ws3.append(["日期", "客户", "事由", "类型", "分类", "金额（元）", "对方", "备注"])
    for cell in ws3[3]:
        cell.font = head_font
        cell.fill = head_fill
    for t in txns:
        ws3.append([
            (t["created_at"] or "")[:10], t.get("customer_name") or "",
            t["item"], "收入" if t["trans_type"] == "income" else "支出",
            t["friendly"], t["amount"], t.get("counterparty") or "", t.get("note") or "",
        ])
    for row in ws3.iter_rows(min_row=4, max_row=ws3.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border

    # 列宽（跳过合并单元格）
    from openpyxl.utils import get_column_letter
    for ws_ in (ws, ws2, ws3):
        for col_idx in range(1, ws_.max_column + 1):
            max_len = 0
            for row in ws_.iter_rows(min_col=col_idx, max_col=col_idx):
                v = row[0].value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws_.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 40)

    out_dir = Path(out_dir) if out_dir else DATA_DIR / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"收支报表_{year}年{month}月.xlsx"
    # Pattern 21: Atomic Write — 先保存 .tmp 再 os.replace，防止报表写到一半损坏
    tmp = out_path.with_suffix(out_path.suffix + ".tmp")
    try:
        wb.save(str(tmp))
        os.replace(str(tmp), str(out_path))
    except Exception:
        tmp.unlink(missing_ok=True)
        raise
    return {"status": "ok", "file": str(out_path), "year": year, "month": month}